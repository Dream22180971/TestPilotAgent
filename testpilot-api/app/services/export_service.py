"""Export service: generate downloadable files in Excel, JSON, or Markdown format."""

from __future__ import annotations

import io
import json
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy.orm import Session

from app.services.db_store import DatabaseStore

store = DatabaseStore()

STORE_TARGETS = ["test_strategy", "test_plan", "test_points", "test_cases", "test_script"]
TARGET_LABELS = {
    "test_strategy": "测试策略",
    "test_plan": "测试计划",
    "test_points": "测试点",
    "test_cases": "测试用例",
    "test_script": "测试脚本",
}


def export_outputs(db: Session, project_id: str, export_format: str, targets: list[str]) -> Any:
    """Export specified outputs for a project in the given format.

    Returns a dict with either file content (as bytes in a dict for JSON response)
    or triggers a StreamingResponse from the route layer.
    """
    outputs: dict[str, Any] = {}
    for t in targets:
        if t in STORE_TARGETS:
            out = store.get_output(db, project_id, t)
            if out:
                outputs[t] = out["content"]

    if export_format == "json":
        return _export_json(project_id, outputs)
    elif export_format == "markdown" or export_format == "md":
        return _export_markdown(project_id, outputs)
    elif export_format == "xlsx" or export_format == "excel":
        return _export_excel(project_id, outputs)
    else:
        return {"project_id": project_id, "format": export_format, "targets": targets, "status": "unsupported_format"}


def _export_json(project_id: str, outputs: dict[str, Any]) -> dict:
    return {
        "project_id": project_id,
        "format": "json",
        "status": "ready",
        "data": outputs,
    }


def _export_markdown(project_id: str, outputs: dict[str, Any]) -> dict:
    lines = [f"# TestPilot 测试输出 - {project_id}\n"]

    for target_key, content in outputs.items():
        label = TARGET_LABELS.get(target_key, target_key)
        lines.append(f"## {label}\n")
        lines.append(_dict_to_md(content))
        lines.append("")

    return {
        "project_id": project_id,
        "format": "markdown",
        "status": "ready",
        "data": "\n".join(lines),
    }


def _export_excel(project_id: str, outputs: dict[str, Any]) -> dict:
    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="101828", end_color="101828", fill_type="solid")

    for target_key, content in outputs.items():
        label = TARGET_LABELS.get(target_key, target_key)
        ws = wb.create_sheet(title=label[:31])  # Excel sheet name max 31 chars

        if target_key == "test_points":
            rows = content.get("rows", [])
            if rows:
                headers = list(rows[0].keys())
                _write_excel_rows(ws, headers, rows, header_font, header_fill)
        elif target_key == "test_cases":
            cases = content.get("cases", [])
            if cases:
                headers = list(cases[0].keys())
                _write_excel_rows(ws, headers, cases, header_font, header_fill)
        elif target_key == "test_strategy":
            _write_dict_sheet(ws, content, header_font, header_fill)
        elif target_key == "test_plan":
            _write_dict_sheet(ws, content, header_font, header_fill)
        elif target_key == "test_script":
            _write_dict_sheet(ws, content, header_font, header_fill)
        else:
            _write_dict_sheet(ws, content, header_font, header_fill)

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return {
        "project_id": project_id,
        "format": "xlsx",
        "status": "ready",
        "data": buf.getvalue().hex(),
        "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }


def _write_excel_rows(ws, headers, rows, header_font, header_fill):
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            val = row.get(header, "")
            if isinstance(val, list):
                val = "; ".join(str(v) for v in val)
            ws.cell(row=row_idx, column=col_idx, value=val)


def _write_dict_sheet(ws, content, header_font, header_fill):
    items = list(content.items())
    for col_idx, (key, val) in enumerate(items, start=1):
        cell = ws.cell(row=1, column=col_idx, value=key)
        cell.font = header_font
        cell.fill = header_fill
        if isinstance(val, list):
            val = "; ".join(str(v) for v in val)
        elif isinstance(val, dict):
            val = json.dumps(val, ensure_ascii=False)
        ws.cell(row=2, column=col_idx, value=str(val))


def _dict_to_md(data: Any, indent: int = 0) -> str:
    prefix = "  " * indent
    if isinstance(data, dict):
        lines = []
        for key, val in data.items():
            if isinstance(val, (dict, list)):
                lines.append(f"{prefix}- **{key}**:")
                lines.append(_dict_to_md(val, indent + 1))
            else:
                lines.append(f"{prefix}- **{key}**: {val}")
        return "\n".join(lines)
    elif isinstance(data, list):
        lines = []
        for item in data:
            if isinstance(item, (dict, list)):
                lines.append(_dict_to_md(item, indent))
            else:
                lines.append(f"{prefix}- {item}")
        return "\n".join(lines)
    else:
        return f"{prefix}{data}"
